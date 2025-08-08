import fitz  # PyMuPDF
import pandas as pd
from tkinter import Tk, filedialog, simpledialog, messagebox
import os
import re
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def parser_secuencial(reg):
    reg = str(reg).strip()
    i = 0
    L = len(reg)
    exp = ''
    guion_count = 0
    while i < L and guion_count < 2:
        exp += reg[i]
        if reg[i] == '-':
            guion_count += 1
        i += 1
    while i < L and (reg[i].isalnum()):
        exp += reg[i]
        i += 1
    nombre = ''
    while i < L and (reg[i].isalpha() or reg[i].isspace()):
        nombre += reg[i]
        i += 1
    nombre = nombre.strip().title()
    rut = ''
    while i < L and reg[i].isdigit():
        rut += reg[i]
        i += 1
    if i < L and reg[i] == '-':
        rut += reg[i]
        i += 1
        if i < L and (reg[i].isdigit() or reg[i].upper() == 'K' or reg[i].isalpha()):
            rut += reg[i]
            i += 1
    while i < L and not reg[i].isdigit():
        i += 1
    fecha = ''
    count_fecha = 0
    while i < L and count_fecha < 10:
        fecha += reg[i]
        i += 1
        count_fecha += 1
    while i < L and not (reg[i].isalpha() or reg[i].isspace()):
        i += 1
    beneficio = ''
    while i < L and (reg[i].isalpha() or reg[i].isspace()):
        beneficio += reg[i]
        i += 1
    beneficio = beneficio.strip().title()
    while i < L and not (reg[i].isdigit() or reg[i] == '.'):
        i += 1
    monto = ''
    while i < L and (reg[i].isdigit() or reg[i] == '.'):
        monto += reg[i]
        i += 1
    sucursal = ''
    while i < L:
        if reg[i].isalpha() or reg[i].isspace():
            sucursal += reg[i]
        i += 1
    sucursal = sucursal.strip().title()
    monto_num = int(monto.replace('.', '')) if monto else 0
    return {
        'expediente': exp.strip(),
        'nombre': nombre,
        'rut': rut.strip(),
        'fecha': fecha.strip(),
        'beneficio': beneficio,
        'monto': monto_num,
        'sucursal': sucursal
    }

def cerrar_excel_si_abierto(archivo_xlsx):
    import win32com.client
    from time import sleep
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        for wb in xl.Workbooks:
            if wb.FullName.lower() == os.path.abspath(archivo_xlsx).lower():
                wb.Close(False)
                sleep(1)
                break
        xl.Quit()
    except:
        pass

def abrir_excel(ruta):
    try:
        os.startfile(ruta)
    except Exception as e:
        print("No se pudo abrir el archivo:", e)

def main():
    root = Tk()
    root.withdraw()

    # Seleccionar PDF
    ruta_pdf = filedialog.askopenfilename(title="Selecciona el PDF", filetypes=[("PDF files", "*.pdf")])
    if not ruta_pdf:
        return

    # Crear carpeta destino
    carpeta_destino = os.path.join(os.path.expanduser("~"), "OneDrive - Chileatiende", "Escritorio", "GUIAS_EN_ACCESS")
    os.makedirs(carpeta_destino, exist_ok=True)

    nombre_base = os.path.splitext(os.path.basename(ruta_pdf))[0]
    ruta_txt = os.path.join(carpeta_destino, nombre_base + '_extraido.txt')
    ruta_xlsx_final = os.path.join(carpeta_destino, f"{nombre_base}_registros.xlsx")

    # Extraer texto del PDF y guardar .txt
    with fitz.open(ruta_pdf) as pdf:
        texto = ""
        for pagina in pdf:
            texto += pagina.get_text()
    with open(ruta_txt, 'w', encoding='utf-8') as f:
        f.write(texto)

    # Leer número de guía desde segunda línea del TXT
    with open(ruta_txt, 'r', encoding='utf-8') as f:
        lineas_txt = f.readlines()
    guia_extraida = ''
    if len(lineas_txt) >= 2:
        match = re.search(r'GUIA\s*:\s*(\d+)', lineas_txt[1].upper())
        if match:
            guia_extraida = match.group(1)

    # Confirmar número de guía
    while True:
        nro_guia = simpledialog.askstring("Número de Guía", "Ingrese o confirme el número de guía:", initialvalue=guia_extraida)
        if nro_guia is None:
            return
        try:
            nro_guia_num = int(nro_guia)
            break
        except ValueError:
            messagebox.showerror("Error", "Debe ingresar solo números.")

    # Leer el texto en binario para limpiar registros
    with open(ruta_txt, 'rb') as f:
        contenido = f.read()
    for cod in ['utf-8', 'latin1', 'cp1252']:
        try:
            texto = contenido.decode(cod)
            break
        except:
            continue
    else:
        texto = contenido.decode('utf-8', errors='replace')

    def limpiar_linea(linea):
        return ''.join([c if c.isprintable() or c in '\t ' else ' ' for c in linea])
    lineas = texto.splitlines()
    lineas_limpias = [limpiar_linea(l) for l in lineas]
    patron = re.compile(r'^\s*\d{2}-\d{7,8}-[A-Z0-9]')
    registros = []
    grabando = False
    for l in lineas_limpias:
        if patron.match(l):
            grabando = True
        if grabando:
            if l.strip().startswith('TOTAL'):
                break
            if l.strip():
                registros.append(l.strip())

    resultados = [parser_secuencial(r) for r in registros]
    for r in resultados:
        r['nro_guia'] = nro_guia_num
    df = pd.DataFrame(resultados)

    # Cerrar si Excel está abierto
    cerrar_excel_si_abierto(ruta_xlsx_final)

    # Guardar archivo Excel sin formato primero
    df.to_excel(ruta_xlsx_final, index=False)

    # Aplicar formato tabla
    wb = load_workbook(ruta_xlsx_final)
    ws = wb.active
    ultima_fila = ws.max_row

    tabla = Table(displayName="TablaRegistros", ref=f"A1:H{ultima_fila}")
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    # Formato numérico columnas
    col_monto = df.columns.get_loc('monto') + 1
    col_guia = df.columns.get_loc('nro_guia') + 1
    for col in ws.iter_cols(min_col=col_monto, max_col=col_monto, min_row=2, max_row=ultima_fila):
        for cell in col:
            cell.number_format = '#,##0'
    for col in ws.iter_cols(min_col=col_guia, max_col=col_guia, min_row=2, max_row=ultima_fila):
        for cell in col:
            cell.number_format = '0'

    wb.save(ruta_xlsx_final)

    # Abrir Excel
    abrir_excel(ruta_xlsx_final)

#if __name__ == "__main__":
 #   main()
if __name__ == "__main__":
    import sys
    import multiprocessing
    multiprocessing.freeze_support()  # Para compatibilidad con Windows
    main()